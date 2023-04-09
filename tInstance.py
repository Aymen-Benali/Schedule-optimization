from tkinter.filedialog import askopenfile
import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook
import random 

State=None
Job=None
num_mach=None
PT=[]
Machine=[]
mach_name=[]

def Generate():
    file = askopenfile(mode ='r', filetypes =[('Excel Files', '*.xlsx *.xlsm *.sxc *.ods *.csv *.tsv')])
    book = load_workbook(filename = file.name, data_only=True ) # Load into openpyxl
    sheet1=book.active
    global Job  
    global State
    global Machine
    global PT
    global num_mach
    global mach_name
    Job = sheet1.max_row-3 #jobs
    State = sheet1.max_column-2
    Machine=[]

    for i in range(State):
        Machine.append(1)
    num_mach=0
    for i in range(len(Machine)):
        num_mach+=Machine[i]
    PT=[]
    p=np.array([])
    p= np.empty((State,Job),int)
    for i in range(3,State+3):
        mach_name.append(sheet1.cell(row=2,column=i).value)
    print(mach_name)
    for i in range(3,State+3 ):
        # for j in range(Machine[i]):
        for k in range(3,Job+3):
            p[i-3,k-3]=sheet1.cell(row=k,column=i).value 
    
    for i in range(State):
        Si=[]
        for j in range (Machine[i]):
            S0=[p[i,k] for k in range(Job)]
            Si.append(S0)
            PT.append(Si)
    print(num_mach)
    print(Job)
    print(State)
    print(PT)
    return PT

PT=Generate()

